#!/usr/bin/env python3
"""
Music League Weekly Updater
Run after scraping new round data to update the xlsx and HTML.

Usage:
  python3 update_music_league.py --new-songs '[ {"league":"...", "round":"...", ...}, ... ]'

Or call update_from_list(new_songs) from another script.
"""

import json, sys, os, re
from pathlib import Path

BASE = Path(__file__).parent
XLSX = BASE / "Random_Thunder_Music_League.xlsx"
VOTES_XLSX = BASE / "Random_Thunder_Music_League_Votes.xlsx"
HTML = BASE / "music_league.html"
VOTES_HTML = BASE / "music_league_votes.html"
ANALYTICS_JSON = BASE / "rtml_analytics.json"


def update_from_list(new_songs, analytics=None):
    """
    new_songs: list of dicts with keys:
      league, round, roundNum, song, artist, album, score, place, submitter
    analytics (optional): dict with keys:
      players, playerStats, roundSummaries, songVotes
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb.active

    # Build set of existing (league, round, song, artist) to avoid dupes
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        league, rnd, rnum, song, artist = row[0], row[1], row[2], row[3], row[4]
        existing.add((str(league), str(rnd), str(song), str(artist)))

    added = 0
    for s in new_songs:
        key = (str(s.get('league','')), str(s.get('round','')), str(s.get('song','')), str(s.get('artist','')))
        if key in existing:
            continue
        ws.append([
            s.get('league',''), s.get('round',''), s.get('roundNum',0),
            s.get('song',''), s.get('artist',''), s.get('album',''),
            s.get('score',0), s.get('place',''), s.get('submitter','')
        ])
        existing.add(key)
        added += 1

    wb.save(str(XLSX))
    print(f"Added {added} new songs to {XLSX}")

    if analytics:
        _save_analytics(analytics)

    if added > 0 or analytics:
        regenerate_html()
    return added


def _save_analytics(analytics):
    """Merge new analytics data into rtml_analytics.json."""
    existing = {}
    if ANALYTICS_JSON.exists():
        existing = json.loads(ANALYTICS_JSON.read_text(encoding='utf-8'))
    existing.update(analytics)
    ANALYTICS_JSON.write_text(json.dumps(existing, ensure_ascii=False), encoding='utf-8')
    print(f"Analytics JSON updated")


def _build_songs_js(songs):
    """Build the compact JS SONGS array string from song list."""
    js_lines = []
    for s in songs:
        def esc(v):
            if not v: return '""'
            return json.dumps(str(v))
        js_lines.append(
            f'[{esc(s["league"])},{esc(s["round"])},{s["roundNum"]},'
            f'{esc(s["song"])},{esc(s["artist"])},{esc(s["album"])},'
            f'{esc(s["place"])},{esc(s["submitter"])}]'
        )
    return "const SONGS=[" + ",".join(js_lines) + "];"


def regenerate_html():
    """Re-embed the full song database and analytics into both HTML files."""
    import openpyxl
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb.active

    songs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        league, round_name, round_num, song, artist, album, score, place, submitter = row
        songs.append({
            "league": league or "",
            "round": round_name or "",
            "roundNum": round_num or 0,
            "song": song or "",
            "artist": artist or "",
            "album": album or "",
            "score": score or 0,
            "place": place or "",
            "submitter": submitter or ""
        })

    new_data = _build_songs_js(songs)

    # Stats
    unique_rounds = len(set((s["league"], s["round"]) for s in songs))
    unique_players = len(set(s["submitter"] for s in songs if s["submitter"]))
    seasons = len(set(s["league"] for s in songs))
    total = len(songs)

    # ── Regenerate music_league.html + index.html ──────────────────────────
    html = HTML.read_text(encoding="utf-8")
    html = re.sub(r'const SONGS=\[.*?\];', lambda _: new_data, html, flags=re.DOTALL)
    html = re.sub(r"countUp\(document\.getElementById\('stat-songs'\),\d+,",
                  f"countUp(document.getElementById('stat-songs'),{total},", html)
    html = re.sub(r"countUp\(document\.getElementById\('stat-rounds'\),\d+,",
                  f"countUp(document.getElementById('stat-rounds'),{unique_rounds},", html)
    html = re.sub(r"countUp\(document\.getElementById\('stat-seasons'\),\d+,",
                  f"countUp(document.getElementById('stat-seasons'),{seasons},", html)
    html = re.sub(r"countUp\(document\.getElementById\('stat-players'\),\d+,",
                  f"countUp(document.getElementById('stat-players'),{unique_players},", html)
    HTML.write_text(html, encoding="utf-8")
    print(f"music_league.html regenerated: {total} songs, {unique_rounds} rounds, {seasons} seasons, {unique_players} players")

    # ── Regenerate music_league_votes.html ─────────────────────────────────
    _regenerate_votes_html(songs, total, unique_rounds, seasons, unique_players, new_data)


def _regenerate_votes_html(songs, total, unique_rounds, seasons, unique_players, new_data):
    """Update music_league_votes.html with new SONGS, stats, and analytics."""
    if not VOTES_HTML.exists():
        print("music_league_votes.html not found, skipping")
        return

    html = VOTES_HTML.read_text(encoding="utf-8")

    # Replace SONGS array
    html = re.sub(r'const SONGS=\[.*?\];', lambda _: new_data, html, flags=re.DOTALL)

    # Update countUp stats
    html = re.sub(r"countUp\(document\.getElementById\('stat-songs'\),\d+,",
                  f"countUp(document.getElementById('stat-songs'),{total},", html)
    html = re.sub(r"countUp\(document\.getElementById\('stat-rounds'\),\d+,",
                  f"countUp(document.getElementById('stat-rounds'),{unique_rounds},", html)
    html = re.sub(r"countUp\(document\.getElementById\('stat-seasons'\),\d+,",
                  f"countUp(document.getElementById('stat-seasons'),{seasons},", html)
    html = re.sub(r"countUp\(document\.getElementById\('stat-players'\),\d+,",
                  f"countUp(document.getElementById('stat-players'),{unique_players},", html)

    # Update analytics bar pills
    html = re.sub(r'<span class="abar-pill">\d+ Seasons</span>',
                  f'<span class="abar-pill">{seasons} Seasons</span>', html)
    html = re.sub(r'<span class="abar-pill">\d+ Rounds</span>',
                  f'<span class="abar-pill">{unique_rounds} Rounds</span>', html)
    html = re.sub(r'<span class="abar-pill">[\d,]+ Songs</span>',
                  f'<span class="abar-pill">{total:,} Songs</span>', html)

    # Update result-count span (static fallback)
    html = re.sub(r'<span id="result-count">[\d,]+</span>',
                  f'<span id="result-count">{total:,}</span>', html)

    # If analytics JSON has updated data, re-embed PLAYERS and PLAYER_STATS
    if ANALYTICS_JSON.exists():
        analytics = json.loads(ANALYTICS_JSON.read_text(encoding="utf-8"))

        if analytics.get("players"):
            players_js = "const PLAYERS = " + json.dumps(analytics["players"], ensure_ascii=False) + ";"
            html = re.sub(r'const PLAYERS = \[.*?\];', lambda _: players_js, html, flags=re.DOTALL)

        if analytics.get("playerStats"):
            stats_js = "const PLAYER_STATS = " + json.dumps(analytics["playerStats"], ensure_ascii=False) + ";"
            html = re.sub(r'const PLAYER_STATS = \[.*?\];', lambda _: stats_js, html, flags=re.DOTALL)

        if analytics.get("roundSummaries"):
            rs_js = "const ROUND_SUMMARIES = " + json.dumps(analytics["roundSummaries"], ensure_ascii=False) + ";"
            html = re.sub(r'const ROUND_SUMMARIES = \[.*?\];', lambda _: rs_js, html, flags=re.DOTALL)

        if analytics.get("songVotes"):
            sv_js = "const SONG_VOTES = " + json.dumps(analytics["songVotes"], ensure_ascii=False) + ";"
            html = re.sub(r'const SONG_VOTES = \[.*?\];', lambda _: sv_js, html, flags=re.DOTALL)

    VOTES_HTML.write_text(html, encoding="utf-8")
    # index.html served by GitHub Pages — keep it in sync with the votes site
    (BASE / "index.html").write_text(html, encoding="utf-8")
    print(f"music_league_votes.html + index.html regenerated: {total} songs, {unique_rounds} rounds, {seasons} seasons, {unique_players} players")


if __name__ == "__main__":
    if '--new-songs' in sys.argv:
        idx = sys.argv.index('--new-songs')
        data = json.loads(sys.argv[idx+1])
        analytics = None
        if '--analytics' in sys.argv:
            aidx = sys.argv.index('--analytics')
            analytics = json.loads(sys.argv[aidx+1])
        update_from_list(data, analytics)
    elif '--regenerate' in sys.argv:
        regenerate_html()
    else:
        print("Usage:")
        print("  python3 update_music_league.py --new-songs '[...]' --analytics '{...}'")
        print("  python3 update_music_league.py --regenerate")
